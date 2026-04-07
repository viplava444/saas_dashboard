"""
scrubber/output_generator.py
────────────────────────────
Writes QC results to:
  • A formatted Excel workbook (one sheet per report)
  • A cleaned CSV of the GC-filtered dataset
  • A JSON summary report

All behaviour is governed by the ``output`` config block.
The Excel formatting (colour-coded duplicates, yellow speeders, bold
headers, auto-sized columns) is identical to the original script but
now fully driven by config values and decoupled from report logic.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scrubber.qc_engine import QCResult

logger = logging.getLogger(__name__)


class OutputGenerator:
    """
    Writes QC results to disk.

    Parameters
    ----------
    cfg:
        The full merged config dict.
    """

    def __init__(self, cfg: dict) -> None:
        self._cfg    = cfg
        self._out    = cfg["output"]
        self._colors = cfg.get("duplicate_colors", [
            "FFB6C1", "FFD700", "98FB98", "87CEEB", "DDA0DD",
        ])
        self._out_dir = Path(self._out.get("output_dir", "output"))
        self._out_dir.mkdir(parents=True, exist_ok=True)

    # ── Public ─────────────────────────────────────────────────────────

    def write(self, result: QCResult, gc_filtered_df: pd.DataFrame) -> None:
        """Write all configured outputs."""
        if self._out.get("excel", True):
            self._write_excel(result)

        if self._out.get("csv", True):
            self._write_csv(gc_filtered_df)

        self._write_summary(result.summary)

    # ── Excel ──────────────────────────────────────────────────────────

    def _write_excel(self, result: QCResult) -> None:
        path = self._out_dir / self._out.get("excel_filename", "survey_QC_report.xlsx")

        sheets: dict[str, pd.DataFrame] = {}

        if not result.ip_duplicates.empty:
            sheets["Duplicate IP"] = result.ip_duplicates

        if not result.identifier_duplicates.empty:
            sheets["Duplicate identifier"] = result.identifier_duplicates

        for label, rpt in result.speeders.items():
            if not rpt.empty:
                sheets[label] = rpt

        if not result.oe_flags.empty:
            sheets["OE Flags"] = result.oe_flags

        if not sheets:
            logger.info("No QC issues found — Excel report not written.")
            return

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply formatting in a second pass (openpyxl after pandas close)
        wb = load_workbook(path)

        if "Duplicate IP" in wb.sheetnames:
            self._fmt_duplicate_sheet(
                wb["Duplicate IP"], result.ip_duplicates, "Duplicate_IP"
            )
        if "Duplicate identifier" in wb.sheetnames:
            self._fmt_duplicate_sheet(
                wb["Duplicate identifier"],
                result.identifier_duplicates,
                "Duplicate_Identifier",
            )
        for label in result.speeders:
            if label in wb.sheetnames:
                self._fmt_speeder_sheet(wb[label], result.speeders[label])

        if "OE Flags" in wb.sheetnames:
            self._fmt_oe_sheet(wb["OE Flags"], result.oe_flags)

        wb.save(path)
        logger.info("Excel report saved: %s", path)

    # ── Formatters ─────────────────────────────────────────────────────

    def _fmt_duplicate_sheet(self, ws, report_df: pd.DataFrame, value_col: str) -> None:
        """Colour-code each unique duplicate value with a distinct colour."""
        if report_df.empty:
            return

        unique_vals = report_df[value_col].unique()
        color_map = {
            val: self._colors[i % len(self._colors)]
            for i, val in enumerate(unique_vals)
        }

        for row_idx, (_, row) in enumerate(report_df.iterrows(), start=2):
            hex_color = color_map[row[value_col]]
            fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            for col_letter in ("A", "B"):
                ws[f"{col_letter}{row_idx}"].fill = fill

        self._bold_header(ws)
        self._auto_size(ws)

    def _fmt_speeder_sheet(self, ws, report_df: pd.DataFrame) -> None:
        """Highlight all speeder rows in yellow."""
        if report_df.empty:
            return

        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row_idx in range(2, len(report_df) + 2):
            for col_letter in ("A", "B"):
                ws[f"{col_letter}{row_idx}"].fill = yellow

        self._bold_header(ws)
        self._auto_size(ws)

    def _fmt_oe_sheet(self, ws, report_df: pd.DataFrame) -> None:
        """Light-orange highlight for OE flagged rows."""
        if report_df.empty:
            return

        orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
        for row_idx in range(2, len(report_df) + 2):
            for col_idx in range(1, len(report_df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = orange

        self._bold_header(ws)
        self._auto_size(ws)

    # ── Shared formatting helpers ──────────────────────────────────────

    @staticmethod
    def _bold_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)

    @staticmethod
    def _auto_size(ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # ── CSV ────────────────────────────────────────────────────────────

    def _write_csv(self, df: pd.DataFrame) -> None:
        path = self._out_dir / self._out.get("csv_filename", "survey_cleaned.csv")
        df.to_csv(path, index=False, encoding="utf-8-sig")
        logger.info("Cleaned CSV saved: %s", path)

    # ── Summary JSON ───────────────────────────────────────────────────

    def _write_summary(self, summary: dict) -> None:
        path = self._out_dir / self._out.get("summary_filename", "qc_summary.json")
        with path.open("w", encoding="utf-8") as fh:
            json.dump(summary, fh, indent=2)
        logger.info("QC summary saved: %s", path)
